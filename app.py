import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import io, openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment,  
                              Border, Side, numbers)
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════════════
# TABEL REFERENSI PKJI 2023
# ══════════════════════════════════════════════════════════════════
# Tabel 4-1: C0 — 4/2-T & 6/2-T per lajur satu arah (×jumlah lajur)
CO  = {"2/2-TT": 2800, "4/2-T": 1700*2, "6/2-T": 1700*3}
# Tabel 4-12: vBD MP (4/2-T, 6/2-T, 8/2-T dalam satu baris = 61)
VBD = {"2/2-TT": 44, "4/2-T": 61, "6/2-T": 61}
# Tabel 4-13: vBL
VBL_T  = {3.00:-4, 3.25:-2, 3.50:0, 3.75:2, 4.00:4}
VBL_TT = {5:-9.5, 6:-3, 7:0, 8:3, 9:4, 10:6, 11:7}
# Tabel 4-3: FCLJ
FCLJ_T  = {3.00:0.92, 3.25:0.96, 3.50:1.00, 3.75:1.04, 4.00:1.08}
FCLJ_TT = {5:0.56, 6:0.87, 7:1.00, 8:1.14, 9:1.25, 10:1.29, 11:1.34}
# Tabel 4-4: FCPA
FCPA = {50:1.00, 55:0.97, 60:0.94, 65:0.91, 70:0.88}
# Tabel 4-14: FVBHS berbahu
FVBHS_BAHU = {
    "4/2-T": {"SR":{0.5:1.02,1.0:1.03,1.5:1.03,2.0:1.04},
              "R": {0.5:0.98,1.0:1.00,1.5:1.02,2.0:1.03},
              "S": {0.5:0.94,1.0:0.97,1.5:1.00,2.0:1.02},
              "T": {0.5:0.89,1.0:0.93,1.5:0.96,2.0:0.99},
              "ST":{0.5:0.84,1.0:0.88,1.5:0.92,2.0:0.96}},
    "2/2-TT":{"SR":{0.5:1.00,1.0:1.01,1.5:1.01,2.0:1.01},
              "R": {0.5:0.96,1.0:0.98,1.5:0.99,2.0:1.00},
              "S": {0.5:0.90,1.0:0.93,1.5:0.96,2.0:0.99},
              "T": {0.5:0.82,1.0:0.86,1.5:0.90,2.0:0.95},
              "ST":{0.5:0.73,1.0:0.79,1.5:0.85,2.0:0.91}},
}
# Tabel 4-15: FVBHS berkereb
FVBHS_KEREB = {
    "4/2-T": {"SR":{0.5:1.00,1.0:1.01,1.5:1.01,2.0:1.02},
              "R": {0.5:0.97,1.0:0.98,1.5:0.99,2.0:1.00},
              "S": {0.5:0.93,1.0:0.95,1.5:0.97,2.0:0.99},
              "T": {0.5:0.87,1.0:0.90,1.5:0.93,2.0:0.96},
              "ST":{0.5:0.81,1.0:0.85,1.5:0.88,2.0:0.92}},
    "2/2-TT":{"SR":{0.5:0.98,1.0:0.99,1.5:0.99,2.0:1.00},
              "R": {0.5:0.93,1.0:0.95,1.5:0.96,2.0:0.98},
              "S": {0.5:0.87,1.0:0.89,1.5:0.92,2.0:0.95},
              "T": {0.5:0.78,1.0:0.81,1.5:0.84,2.0:0.88},
              "ST":{0.5:0.68,1.0:0.72,1.5:0.77,2.0:0.82}},
}
# Tabel 4-16: FVBUK
FVBUK_T = {"< 0,1 juta":0.90,"0,1 – 0,5 juta":0.93,
            "0,5 – 1,0 juta":0.95,"1,0 – 3,0 juta":1.00,"> 3,0 juta":1.03}
# Tabel 4-5: FCHS berbahu
FCHS_BAHU = {
    "4/2-T": {"SR":{0.5:0.96,1.0:0.98,1.5:1.01,2.0:1.03},
              "R": {0.5:0.94,1.0:0.97,1.5:1.00,2.0:1.02},
              "S": {0.5:0.92,1.0:0.95,1.5:0.98,2.0:1.00},
              "T": {0.5:0.88,1.0:0.92,1.5:0.95,2.0:0.98},
              "ST":{0.5:0.84,1.0:0.88,1.5:0.92,2.0:0.96}},
    "2/2-TT":{"SR":{0.5:0.94,1.0:0.96,1.5:0.99,2.0:1.01},
              "R": {0.5:0.92,1.0:0.94,1.5:0.97,2.0:1.00},
              "S": {0.5:0.89,1.0:0.92,1.5:0.95,2.0:0.98},
              "T": {0.5:0.82,1.0:0.86,1.5:0.90,2.0:0.95},
              "ST":{0.5:0.73,1.0:0.79,1.5:0.85,2.0:0.91}},
}
# Tabel 4-6: FCHS berkereb
FCHS_KEREB = {
    "4/2-T": {"SR":{0.5:0.95,1.0:0.97,1.5:0.99,2.0:1.01},
              "R": {0.5:0.94,1.0:0.96,1.5:0.98,2.0:1.00},
              "S": {0.5:0.91,1.0:0.93,1.5:0.95,2.0:0.98},
              "T": {0.5:0.86,1.0:0.89,1.5:0.92,2.0:0.95},
              "ST":{0.5:0.81,1.0:0.85,1.5:0.88,2.0:0.92}},
    "2/2-TT":{"SR":{0.5:0.93,1.0:0.95,1.5:0.97,2.0:0.99},
              "R": {0.5:0.90,1.0:0.92,1.5:0.95,2.0:0.97},
              "S": {0.5:0.86,1.0:0.88,1.5:0.91,2.0:0.94},
              "T": {0.5:0.78,1.0:0.81,1.5:0.84,2.0:0.88},
              "ST":{0.5:0.68,1.0:0.72,1.5:0.77,2.0:0.82}},
}
# Tabel 4-7: FCUK
FCUK_T = {"< 0,1 juta":0.86,"0,1 – 0,5 juta":0.90,
           "0,5 – 1,0 juta":0.94,"1,0 – 3,0 juta":1.00,"> 3,0 juta":1.04}
# Gambar 4-1 (2/2-TT) & 4-2 (terbagi): rumus polinomial vT = f(DJ) per kelas vB
# (hasil regresi kurva, menggantikan pembacaan titik+interpolasi)
def _vt_22tt_30(dj):
    return (-5.4356*dj**5)-(7.3788*dj**4)+(18.358*dj**3)-(8.9749*dj**2)-(8.7817*dj)+29.985
def _vt_22tt_40(dj):
    return (-37.91*dj**5)+(58*dj**4)-(24.843*dj**3)+(0.7485*dj**2)-(12.598*dj)+39.949
def _vt_22tt_50(dj):
    return (-114.09*dj**5)+(238.5*dj**4)-(178.49*dj**3)+(55.425*dj**2)-(22.523*dj)+50.08
def _vt_22tt_60(dj):
    return (65.539*dj**5)-(175.99*dj**4)+(153.51*dj**3)-(51.563*dj**2)-(14.391*dj)+59.951
def _vt_22tt_70(dj):
    return (93.528*dj**5)-(268.17*dj**4)+(255.73*dj**3)-(97.001*dj**2)-(10.813*dj)+69.943

def _vt_terbagi_40(dj):
    return 40-2*dj-9*dj**2
def _vt_terbagi_50(dj):
    return (-174.97*dj**5)+(372.77*dj**4)-(277.37*dj**3)+(70.096*dj**2)-(10.286*dj)+50.138
def _vt_terbagi_60(dj):
    return (-320.9*dj**5)+(697.12*dj**4)-(535.4*dj**3)+(155.55*dj**2)-(21.56*dj)+60.141
def _vt_terbagi_70(dj):
    return (-310.7*dj**5)+(660.54*dj**4)-(492.75*dj**3)+(132.95*dj**2)-(18.888*dj)+70.124
def _vt_terbagi_80(dj):
    return (-96.126*dj**5)+(138.88*dj**4)-(50.019*dj**3)-(22.106*dj**2)-(1.7704*dj)+80.111

RUMUS_VT_22TT    = {30:_vt_22tt_30, 40:_vt_22tt_40, 50:_vt_22tt_50,
                     60:_vt_22tt_60, 70:_vt_22tt_70}
RUMUS_VT_TERBAGI = {40:_vt_terbagi_40, 50:_vt_terbagi_50, 60:_vt_terbagi_60,
                     70:_vt_terbagi_70, 80:_vt_terbagi_80}

# ══════════════════════════════════════════════════════════════════
# FUNGSI INTI
# ══════════════════════════════════════════════════════════════════
def _td(tbl, v):
    return min(tbl.keys(), key=lambda k: abs(k-v))

def _vt_formula(vb, dj, tipe):
    """Evaluasi rumus polinomial vT=f(DJ) (Gambar 4-1/4-2 PKJI 2023) tanpa pembulatan
    akhir — dipakai untuk menggambar kurva halus. vB dibulatkan ke kelas kurva acuan
    terdekat (kelipatan 10: 30/40/50/60/70 untuk 2/2-TT, 40/50/60/70/80 untuk jalan
    terbagi), dan DJ dibatasi maksimum 1,00 karena rumus hanya valid pada rentang kurva
    resmi PKJI 2023 (di luar itu vT ditahan pada nilai DJ=1,00, sama seperti kondisi
    kolaps LOS F pada sistem sebelumnya)."""
    tabel_rumus = RUMUS_VT_22TT if tipe=="2/2-TT" else RUMUS_VT_TERBAGI
    vb_key = _td(tabel_rumus, vb)
    dj_aman = min(max(dj,0.0),1.00)
    return tabel_rumus[vb_key](dj_aman)

def hitung_vt(vb, dj, tipe):
    """Nilai vT akhir (dibulatkan tanpa angka desimal) sesuai rumus polinomial
    Gambar 4-1/4-2 PKJI 2023."""
    return round(_vt_formula(vb, dj, tipe))

def _k6(v): return 1-0.8*(1-v)

def hitung_emp(tipe, vol, lebar):
    if tipe=="2/2-TT":
        ks,sm = (1.3, 0.50 if lebar<6 else 0.40) if vol<1800 \
                else (1.2, 0.35 if lebar<6 else 0.25)
    elif tipe=="4/2-T":
        ks,sm = (1.3,0.40) if vol/2<1050 else (1.2,0.25)
    else:
        ks,sm = (1.3,0.40) if vol/3<1100 else (1.2,0.25)
    return {"KR":1.0,"KS":ks,"SM":sm}

def _los(dj):
    if   dj<=0.19: return "A","Arus bebas, kepadatan sangat rendah","success"
    elif dj<=0.44: return "B","Arus stabil, kecepatan sedikit terpengaruh","success"
    elif dj<=0.69: return "C","Arus stabil, kecepatan & manuver mulai terbatas","warning"
    elif dj<=0.84: return "D","Arus mendekati tidak stabil, kecepatan menurun","warning"
    elif dj<=1.00: return "E","Arus tidak stabil, kondisi kritis","error"
    else:          return "F","Arus terhambat / macet","error"

def hitung(tipe,lebar,ks,ls,hs,kota,split,kr,kb,sm,pjg):
    vol=kr+kb+sm
    emp=hitung_emp(tipe,vol,lebar)
    q=kr*emp["KR"]+kb*emp["KS"]+sm*emp["SM"]

    vbd=VBD[tipe]
    vbl=VBL_TT[_td(VBL_TT,lebar)] if tipe=="2/2-TT" else VBL_T[_td(VBL_T,lebar)]

    ref="4/2-T" if tipe=="6/2-T" else tipe
    tb_fv=FVBHS_BAHU[ref] if ks=="Berbahu" else FVBHS_KEREB[ref]
    fv4=tb_fv[hs][_td(tb_fv[hs],ls)]
    fvbhs=_k6(fv4) if tipe=="6/2-T" else fv4
    fvbuk=FVBUK_T[kota]
    vb=(vbd+vbl)*fvbhs*fvbuk

    co=CO[tipe]
    fclj=FCLJ_TT[_td(FCLJ_TT,lebar)] if tipe=="2/2-TT" else FCLJ_T[_td(FCLJ_T,lebar)]
    fcpa=FCPA[_td(FCPA,split)] if tipe=="2/2-TT" else 1.00
    tb_fc=FCHS_BAHU[ref] if ks=="Berbahu" else FCHS_KEREB[ref]
    fc4=tb_fc[hs][_td(tb_fc[hs],ls)]
    fchs=_k6(fc4) if tipe=="6/2-T" else fc4
    fcuk=FCUK_T[kota]
    c=co*fclj*fcpa*fchs*fcuk

    dj=q/c if c>0 else 0
    vt=hitung_vt(vb,dj,tipe)
    wt=(pjg/vt*60) if vt>0 else 0
    los,desk,warna=_los(dj)
    return dict(emp=emp,vol=vol,q_kr=kr*emp["KR"],q_kb=kb*emp["KS"],
                q_sm=sm*emp["SM"],q=q,vbd=vbd,vbl=vbl,fv4=fv4,
                fvbhs=fvbhs,fvbuk=fvbuk,vb=vb,co=co,fclj=fclj,
                fcpa=fcpa,fc4=fc4,fchs=fchs,fcuk=fcuk,c=c,
                dj=dj,vt=vt,wt=wt,gambar="4-1" if tipe=="2/2-TT" else "4-2",
                los=los,desk=desk,warna=warna)

# ══════════════════════════════════════════════════════════════════
# EKSPOR EXCEL
# ══════════════════════════════════════════════════════════════════
def warna_los_hex(los):
    return {"A":"#1E8449","B":"#27AE60","C":"#D4AC0D",
            "D":"#E67E22","E":"#E74C3C","F":"#922B21"}.get(los,"#FFFFFF")

def buat_excel(nama,kota_nama,tipe,pjg,ks,ls_label,ls_val,hs,kota,periodes,hasil):
    wb = openpyxl.Workbook()
    # ─── STYLE HELPERS ───────────────────────────────────────────
    def fill(hex_):
        return PatternFill("solid", fgColor=hex_.lstrip("#"))
    def font(bold=False,color="000000",size=11):
        return Font(bold=bold,color=color,size=size)
    def border():
        s=Side(style="thin")
        return Border(left=s,right=s,top=s,bottom=s)
    def center(): return Alignment(horizontal="center",vertical="center",wrap_text=True)
    def wrap():   return Alignment(vertical="center",wrap_text=True)

    BIRU  = "1F3864"; BIRU2 = "2E75B6"; ABU  = "D9E1F2"
    PUTIH = "FFFFFF"; KUNING= "FFF2CC"; HIJAU= "E2EFDA"

    # ─── SHEET 1: REKAPITULASI ───────────────────────────────────
    ws = wb.active
    ws.title = "Rekapitulasi"
    ws.sheet_view.showGridLines = False

    # Judul
    ws.merge_cells("A1:J1")
    ws["A1"] = "LAPORAN ANALISIS KINERJA RUAS JALAN PERKOTAAN"
    ws["A1"].font = font(bold=True,color=PUTIH,size=14)
    ws["A1"].fill = fill(BIRU); ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:J2")
    ws["A2"] = "Berdasarkan PKJI 2023 — Diagram Alir 4-4 (Analisis Operasional)"
    ws["A2"].font = font(color=PUTIH,size=11)
    ws["A2"].fill = fill(BIRU2); ws["A2"].alignment = center()
    ws.row_dimensions[2].height = 20

    # Info ruas
    info = [
        ("Nama Ruas Jalan",  nama),
        ("Kota / Lokasi",    kota_nama),
        ("Tipe Jalan",       tipe),
        ("Panjang Segmen",   f"{pjg} km"),
        ("Kondisi Samping",  f"{ks} ({ls_label} = {ls_val} m)"),
        ("Kelas Ham. Samping", hs),
        ("Ukuran Kota",      kota),
    ]
    for i,(k,v) in enumerate(info,4):
        ws.merge_cells(f"A{i}:D{i}")
        ws[f"A{i}"] = k
        ws[f"A{i}"].font = font(bold=True); ws[f"A{i}"].fill = fill(ABU)
        ws[f"A{i}"].alignment = wrap()
        for col in ["A", "B", "C", "D"] :
            ws[f"{col}{i}"].border = border()
        ws.merge_cells(f"E{i}:J{i}")
        ws[f"E{i}"] = v
        ws[f"E{i}"].alignment = wrap() 
        for col in ["E", "F", "G", "H", "I", "J"] :
            ws[f"{col}{i}"].border = border()
    ws.row_dimensions[3].height = 5

    # Header tabel rekap
    r = len(info)+5
    hdrs = ["No","Periode","Q (smp/jam)","C (smp/jam)","vB (km/jam)",
            "DJ","LOS","vT (km/jam)","wT (menit)","Deskripsi Kinerja"]
    cols = ["A","B","C","D","E","F","G","H","I","J"]
    ws.merge_cells(f"A{r}:J{r}")
    ws[f"A{r}"] = "REKAPITULASI HASIL PERHITUNGAN"
    ws[f"A{r}"].font = font(bold=True,color=PUTIH); ws[f"A{r}"].fill = fill(BIRU2)
    ws[f"A{r}"].alignment = center(); ws.row_dimensions[r].height = 20
    r+=1
    ws.merge_cells(f"J{r}:J{r}")  # gabung kolom deskripsi
    for col,hdr in zip(cols,hdrs):
        c_ = f"{col}{r}" if col!="J" else f"J{r}"
        ws[c_] = hdr
        ws[c_].font = font(bold=True,color=PUTIH)
        ws[c_].fill = fill(BIRU); ws[c_].alignment = center()
        ws[c_].border = border()
    ws.row_dimensions[r].height = 30
    r+=1

    for i,(h,p) in enumerate(zip(hasil,periodes),1):
        hex_los = warna_los_hex(h["los"])
        vals = [i, p["jam"], round(h["q"],2), round(h["c"],2),
                round(h["vb"],2), round(h["dj"],3), h["los"],
                round(h["vt"],2), round(h["wt"],2), h["desk"]]
        for col,val in zip(cols,vals):
            cell = ws[f"{col}{r}"] if col!="J" else ws[f"J{r}"]
            cell.value = val
            cell.border = border()
            cell.alignment = center() if col!="J" else wrap()
            if col=="F":  # DJ — warna sesuai LOS
                cell.fill = fill(hex_los.lstrip("#") if len(hex_los)==7 else hex_los)
                cell.font = font(bold=True,color=PUTIH)
            elif col=="G":  # LOS
                cell.fill = fill(hex_los.lstrip("#") if len(hex_los)==7 else hex_los)
                cell.font = font(bold=True,color=PUTIH)
            elif i%2==0:
                cell.fill = fill(ABU)
        ws.row_dimensions[r].height = 22
        r+=1

    # Lebar kolom
    for col,w in zip("ABCDEFGHIJ",[5,20,14,14,14,10,8,14,12,35,3]):
        ws.column_dimensions[col].width = w

    # ─── SHEET per PERIODE ───────────────────────────────────────
    for idx,(h,p) in enumerate(zip(hasil,periodes)):
        ws2 = wb.create_sheet(title=f"Periode {idx+1}")
        ws2.sheet_view.showGridLines = False
        hex_los = warna_los_hex(h["los"])

        # Judul
        ws2.merge_cells("A1:H1")
        ws2["A1"] = f"ANALISIS KINERJA RUAS JALAN — {p['jam']}"
        ws2["A1"].font = font(bold=True,color=PUTIH,size=13)
        ws2["A1"].fill = fill(BIRU); ws2["A1"].alignment = center()
        ws2.row_dimensions[1].height = 28

        ws2.merge_cells("A2:H2")
        ws2["A2"] = f"{nama}  |  Tipe: {tipe}  |  Kota: {kota_nama}"
        ws2["A2"].font = font(color=PUTIH)
        ws2["A2"].fill = fill(BIRU2); ws2["A2"].alignment = center()

        def blok(ws,judul,baris,data_rows,warna_judul=BIRU2):
            ws.merge_cells(f"A{baris}:H{baris}")
            ws[f"A{baris}"] = judul
            ws[f"A{baris}"].font = font(bold=True,color=PUTIH)
            ws[f"A{baris}"].fill = fill(warna_judul)
            ws[f"A{baris}"].alignment = center()
            ws.row_dimensions[baris].height = 18
            baris+=1
            for k,v in data_rows:
                ws.merge_cells(f"A{baris}:C{baris}")
                ws[f"A{baris}"] = k
                ws[f"A{baris}"].font = font(bold=True); ws[f"A{baris}"].fill = fill(ABU)
                for col in ["A", "B", "C"] :
                    ws[f"{col}{baris}"].border = border()
                ws[f"A{baris}"].alignment = wrap()
                ws.merge_cells(f"D{baris}:H{baris}")
                ws[f"D{baris}"] = v
                for col in ["D", "E", "F", "G", "H"] :
                    ws[f"{col}{baris}"].border = border() 
                ws[f"D{baris}"].alignment = wrap()
                ws.row_dimensions[baris].height = 18
                baris+=1
            return baris

        r2 = 4
        # Volume
        r2 = blok(ws2,"A. VOLUME LALU LINTAS",r2,[
            ("KR (kend/jam)",  p["kr"]),
            ("KB (kend/jam)",  p["kb"]),
            ("SM (kend/jam)",  p["sm"]),
            ("EMP KR / KB / SM", f"{h['emp']['KR']:.2f}  /  {h['emp']['KS']:.2f}  /  {h['emp']['SM']:.2f}"),
            ("Q (smp/jam)",    round(h["q"],2)),
        ])
        r2+=1
        # vB
        r2 = blok(ws2,"B. KECEPATAN ARUS BEBAS  vB = (vBD + vBL) × FVBHS × FVBUK",r2,[
            ("vBD (Tabel 4-12)",  f"{h['vbd']} km/jam"),
            ("vBL (Tabel 4-13)",  f"{h['vbl']:+.1f} km/jam"),
            ("FVBHS (Tabel 4-14/4-15)", f"{h['fvbhs']:.3f}" +
             (f"  [FV4HS={h['fv4']:.3f} → Pers.4-5]" if tipe=="6/2-T" else "")),
            ("FVBUK (Tabel 4-16)", f"{h['fvbuk']:.3f}"),
            ("vB (km/jam)",       round(h["vb"],2)),
        ])
        r2+=1
        # C
        r2 = blok(ws2,"C. KAPASITAS  C = C0 × FCLJ × FCPA × FCHS × FCUK",r2,[
            ("C0 (Tabel 4-1)",     f"{h['co']} smp/jam"),
            ("FCLJ (Tabel 4-3)",   f"{h['fclj']:.3f}"),
            ("FCPA (Tabel 4-4)",   f"{h['fcpa']:.3f}"),
            ("FCHS (Tabel 4-5/4-6)",f"{h['fchs']:.3f}" +
             (f"  [FC4HS={h['fc4']:.3f} → Pers.4-2]" if tipe=="6/2-T" else "")),
            ("FCUK (Tabel 4-7)",   f"{h['fcuk']:.3f}"),
            ("C (smp/jam)",        round(h["c"],2)),
        ])
        r2+=1
        # Kinerja
        r2 = blok(ws2,"D. KINERJA LALU LINTAS",r2,[
            ("Derajat Kejenuhan DJ",  round(h["dj"],3)),
            ("Kecepatan Tempuh vT",   f"{h['vt']:.2f} km/jam  (Gambar {h['gambar']})"),
            ("Waktu Tempuh wT",       f"{h['wt']:.2f} menit  (wT = P/vT × 60)"),
        ],warna_judul=BIRU)
        # LOS box
        ws2.merge_cells(f"A{r2}:H{r2}")
        ws2[f"A{r2}"] = f"LOS {h['los']}  —  {h['desk']}"
        ws2[f"A{r2}"].font = font(bold=True,color=PUTIH,size=12)
        ws2[f"A{r2}"].fill = fill(hex_los.lstrip("#"))
        ws2[f"A{r2}"].alignment = center()
        ws2.row_dimensions[r2].height = 24

        for col,w in zip("ABCDEFGH",[6,16,14,14,14,12,10,20]):
            ws2.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════
# GAUGE CHART
# ══════════════════════════════════════════════════════════════════
def gauge(dj, label=""):
    los,_,_ = _los(dj)
    wb = "#27ae60" if dj<=0.44 else "#f39c12" if dj<=0.82 else "#e74c3c"
    fig = go.Figure(go.Indicator(
        mode="gauge+number", value=round(dj,3),
        number={"font":{"size":44,"color":wb}},
        title={"text":f"Derajat Kejenuhan (DJ)<br><b>LOS {los}</b> — {label}",
               "font":{"size":12}},
        gauge={"axis":{"range":[0,1.2],
                       "tickvals":[0,.20,.44,.64,.82,1.00,1.20],
                       "ticktext":["0","0,20","0,44","0,64","0,82","1,00","1,2"],
                       "tickcolor":"white"},
               "bar":{"color":wb,"thickness":0.25},
               "bgcolor":"rgba(0,0,0,0)","borderwidth":0,
               "steps":[{"range":[0,.20],"color":"#1e8449"},
                        {"range":[.20,.44],"color":"#27ae60"},
                        {"range":[.44,.64],"color":"#d4ac0d"},
                        {"range":[.64,.82],"color":"#e67e22"},
                        {"range":[.82,1.00],"color":"#e74c3c"},
                        {"range":[1.00,1.20],"color":"#922b21"}],
               "threshold":{"line":{"color":"white","width":4},
                            "thickness":0.8,"value":min(dj,1.18)}},
    ))
    for x,y,t in [(.10,.18,"A"),(.28,.06,"B"),(.50,.01,"C"),
                  (.67,.06,"D"),(.83,.18,"E"),(.95,.35,"F")]:
        fig.add_annotation(x=x,y=y,text=f"<b>{t}</b>",showarrow=False,
                           font=dict(size=14,color="white"))
    fig.update_layout(height=300,margin=dict(l=30,r=30,t=80,b=10),
                      paper_bgcolor="rgba(0,0,0,0)",font={"color":"white"})
    return fig

# ══════════════════════════════════════════════════════════════════
# KURVA vT vs DJ
# ══════════════════════════════════════════════════════════════════
def kurva_vt(h, tipe, key=""):
    vb_list  = [30,40,50,60,70] if tipe=="2/2-TT" else [40,50,60,70,80]
    judul    = ("Gambar 4-1 — Tipe 2/2-TT" if tipe=="2/2-TT"
                else "Gambar 4-2 — Tipe 4/2-T, 6/2-T, 8/2-T")
    dj_s = [i/100 for i in range(0,71,2)]
    dj_d = [i/100 for i in range(70,103,2)]
    clrs = ["#3498db","#2ecc71","#f39c12","#e74c3c","#9b59b6"]
    fig = go.Figure()
    for vb,c in zip(vb_list,clrs):
        fig.add_trace(go.Scatter(x=dj_s,y=[_vt_formula(vb,d,tipe) for d in dj_s],
            mode="lines",line=dict(color=c,width=2),name=f"vB={vb}",legendgroup=f"{vb}"))
        fig.add_trace(go.Scatter(x=dj_d,y=[_vt_formula(vb,d,tipe) for d in dj_d],
            mode="lines",line=dict(color=c,width=2,dash="dash"),
            showlegend=False,legendgroup=f"{vb}"))
        fig.add_annotation(x=0.01,y=_vt_formula(vb,0.01,tipe),
            text=f"  {vb}",showarrow=False,font=dict(size=10,color=c),xanchor="left")
    # Titik kondisi saat ini
    fig.add_trace(go.Scatter(x=[h["dj"]],y=[h["vt"]],mode="markers",
        marker=dict(color="white",size=14,symbol="star",
                    line=dict(color="red",width=2)),
        name=f"Saat ini (DJ={h['dj']:.3f}, vT={h['vt']:.1f})"))
    fig.add_shape(type="line",x0=h["dj"],x1=h["dj"],y0=0,y1=h["vt"],
                  line=dict(color="white",dash="dot",width=1))
    fig.add_shape(type="line",x0=0,x1=h["dj"],y0=h["vt"],y1=h["vt"],
                  line=dict(color="white",dash="dot",width=1))
    fig.add_annotation(x=h["dj"],y=h["vt"],
        text=f"  DJ={h['dj']:.3f}<br>  vT={h['vt']:.1f}",
        showarrow=False,xanchor="left",
        font=dict(size=10,color="white"),bgcolor="rgba(0,0,0,0.6)")
    for x0,x1,wz in [(0,.20,"#1e8449"),(.20,.44,"#27ae60"),(.44,.64,"#d4ac0d"),
                      (.64,.82,"#e67e22"),(.82,1.00,"#e74c3c")]:
        fig.add_vrect(x0=x0,x1=x1,fillcolor=wz,opacity=0.07,layer="below",line_width=0)
    fig.update_layout(
        title=dict(text=judul,font=dict(size=12)),
        xaxis=dict(title="Derajat Kejenuhan (DJ)",range=[0,1.05],
                   gridcolor="rgba(255,255,255,0.12)"),
        yaxis=dict(title="Kecepatan Tempuh vMP (km/jam)",
                   range=[0,max(vb_list)+10],
                   gridcolor="rgba(255,255,255,0.12)"),
        height=400,
        legend=dict(x=0.62,y=0.98,bgcolor="rgba(0,0,0,0.4)",
                    bordercolor="white",borderwidth=1,font=dict(size=10)),
        paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(20,20,30,0.6)",
        font=dict(color="white"),margin=dict(l=60,r=30,t=50,b=50))
    return fig

# ══════════════════════════════════════════════════════════════════
# ANALISIS SKENARIO  (pakai key unik agar tidak reset halaman)
# ══════════════════════════════════════════════════════════════════
def skenario(tipe,lebar,ks,ls,hs,kota,split,kr,kb,sm,pjg,label,suffix):
    st.subheader("🔮 Analisis Skenario — What If?")
    st.caption(f"Simulasi kenaikan volume dari kondisi eksisting **{label}**.")

    col_sl, col_rd = st.columns([2,1])
    with col_sl:
        maks = st.slider("Rentang kenaikan volume (%)",10,100,50,10,
                         key=f"maks_{suffix}")
    with col_rd:
        langkah = st.radio("Interval",[5,10,20],horizontal=True,
                           key=f"step_{suffix}",
                           format_func=lambda x:f"+{x}%")

    pct_list=[*range(0,maks+langkah,langkah)]
    rows,dj_l,vt_l,lbl_l=[],[],[],[]
    for pct in pct_list:
        f=1+pct/100
        r=hitung(tipe,lebar,ks,ls,hs,kota,split,
                 int(kr*f),int(kb*f),int(sm*f),pjg)
        lbl="Eksisting" if pct==0 else f"+{pct}%"
        rows.append({"Skenario":lbl,
                     "KR":int(kr*f),"KB":int(kb*f),"SM":int(sm*f),
                     "Q (smp/j)":f"{r['q']:.1f}","C (smp/j)":f"{r['c']:.1f}",
                     "DJ":f"{r['dj']:.3f}","LOS":r["los"],
                     "vT (km/j)":f"{r['vt']:.1f}","wT (mnt)":f"{r['wt']:.2f}"})
        dj_l.append(r["dj"]); vt_l.append(r["vt"]); lbl_l.append(lbl)

    st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)

    # Bar DJ
    clr=["#27ae60" if d<=0.44 else "#f39c12" if d<=0.82 else "#e74c3c" for d in dj_l]
    fig=go.Figure()
    fig.add_trace(go.Bar(x=lbl_l,y=dj_l,marker_color=clr,
        text=[f"DJ={d:.3f}<br>LOS {r['LOS']}" for d,r in zip(dj_l,rows)],
        textposition="outside"))
    for bts,lbl_g,wg in [(.44,"B/C","#27ae60"),(.64,"C/D","#d4ac0d"),
                          (.82,"D/E","#e67e22"),(1.00,"E/F","#e74c3c")]:
        fig.add_hline(y=bts,line_dash="dot",line_color=wg,
                      annotation_text=f"Batas {lbl_g}",
                      annotation_position="right",annotation_font_color=wg)
    fig.update_layout(title="Perubahan DJ per Skenario",
        yaxis=dict(range=[0,max(dj_l)*1.35+0.1],title="DJ"),
        xaxis_title="Skenario",height=360,showlegend=False,
        paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="white"))
    st.plotly_chart(fig,use_container_width=True,key=f"bar_{suffix}")

    # Line vT
    fig2=go.Figure()
    fig2.add_trace(go.Scatter(x=lbl_l,y=vt_l,mode="lines+markers+text",
        marker=dict(size=10,color=clr),line=dict(color="white",width=2),
        text=[f"{v:.1f}" for v in vt_l],textposition="top center"))
    fig2.update_layout(title="Perubahan vT per Skenario",
        yaxis=dict(range=[0,max(vt_l)*1.3],title="vT (km/jam)"),
        xaxis_title="Skenario",height=300,
        paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",
        font=dict(color="white"))
    st.plotly_chart(fig2,use_container_width=True,key=f"vt_{suffix}")

    # Peringatan otomatis
    kritis=[(l,d) for l,d in zip(lbl_l,dj_l) if d>0.82]
    jenuh =[(l,d) for l,d in zip(lbl_l,dj_l) if 0.64<d<=0.82]
    if kritis:
        st.error(f"⚠️ LOS E/F mulai pada skenario **{kritis[0][0]}** (DJ={kritis[0][1]:.3f})")
    elif jenuh:
        st.warning(f"⚠️ LOS D mulai pada skenario **{jenuh[0][0]}** (DJ={jenuh[0][1]:.3f})")
    else:
        st.success(f"✅ Semua skenario s.d. +{maks}% masih LOS A–C")

# ══════════════════════════════════════════════════════════════════
# TAMPILAN SATU PERIODE
# ══════════════════════════════════════════════════════════════════
def tampil_periode(h,p,idx,tipe,lebar,ks,ls_label,ls,hs,kota,split,pjg,nama):
    st.plotly_chart(gauge(h["dj"],p["jam"]),
                    use_container_width=True,key=f"g_{idx}_{p['jam']}")
    c1,c2,c3,c4=st.columns(4)
    c1.metric("Volume Q",  f"{h['q']:.0f}","smp/jam")
    c2.metric("Kapasitas C",f"{h['c']:.0f}","smp/jam")
    c3.metric("vT",         f"{h['vt']:.1f}","km/jam")
    c4.metric("wT",         f"{h['wt']:.2f}","menit")
    getattr(st,h["warna"])(f"**DJ = {h['dj']:.3f}  |  LOS {h['los']}**  —  {h['desk']}")
    st.divider()
    t_vb="Tabel 4-14 (bahu)" if ks=="Berbahu" else "Tabel 4-15 (kereb)"
    t_hs="Tabel 4-5 (bahu)"  if ks=="Berbahu" else "Tabel 4-6 (kereb)"
    ca,cb=st.columns(2)
    with ca:
        st.subheader("Kecepatan Arus Bebas (vB)")
        st.table({"Parameter":["vBD","vBL","FVBHS","FVBUK","vB"],
                  "Nilai":[f"{h['vbd']} km/jam",f"{h['vbl']:+.1f} km/jam",
                           f"{h['fvbhs']:.3f}  {t_vb}",f"{h['fvbuk']:.3f}",
                           f"**{h['vb']:.2f} km/jam**"]})
    with cb:
        st.subheader("Kapasitas Jalan (C)")
        st.table({"Parameter":["C0","FCLJ","FCPA","FCHS","FCUK","C"],
                  "Nilai":[f"{h['co']} smp/jam",f"{h['fclj']:.3f}",
                           f"{h['fcpa']:.3f}",f"{h['fchs']:.3f}  {t_hs}",
                           f"{h['fcuk']:.3f}",f"**{h['c']:.2f} smp/jam**"]})
    st.divider()
    st.subheader(f"📈 Kurva vT vs DJ — Gambar {h['gambar']} PKJI 2023")
    st.plotly_chart(kurva_vt(h,tipe,key=f"kv_{idx}"),
                    use_container_width=True,key=f"kurva_{idx}_{p['jam']}")
    st.caption("⭐ = kondisi saat ini  |  Garis penuh = DJ ≤ 0,70  |  Garis putus = DJ > 0,70")
    st.divider()
    st.subheader("EMP (Tabel 4-10/4-11 — dinamis)")
    e1,e2,e3=st.columns(3)
    e1.metric("EMP KR",f"{h['emp']['KR']:.2f}")
    e2.metric("EMP KB",f"{h['emp']['KS']:.2f}")
    e3.metric("EMP SM",f"{h['emp']['SM']:.2f}")
    st.bar_chart(pd.DataFrame({"Jenis":["KR","KB","SM"],
        "kend/jam":[p["kr"],p["kb"],p["sm"]],
        "smp/jam":[round(h["q_kr"],1),round(h["q_kb"],1),round(h["q_sm"],1)],
    }).set_index("Jenis"))
    st.divider()
    skenario(tipe,lebar,ks,ls,hs,kota,split,
             p["kr"],p["kb"],p["sm"],pjg,p["jam"],f"{idx}_{p['jam']}")

# ══════════════════════════════════════════════════════════════════
# SESSION STATE — Simpan hasil agar tidak reset saat widget berubah
# ══════════════════════════════════════════════════════════════════
if "data" not in st.session_state:
    st.session_state.data = None

# ══════════════════════════════════════════════════════════════════
# LAYOUT STREAMLIT
# ══════════════════════════════════════════════════════════════════
st.title("🛣️ Analisis Kinerja Ruas Jalan Perkotaan")
st.write("Berdasarkan **PKJI 2023** — Diagram Alir 4-4 (Analisis Operasional)")
st.divider()

with st.sidebar:
    st.header("📋 Input Data")

    st.subheader("Data Umum")
    nama   = st.text_input("Nama Ruas Jalan","Jl. Contoh")
    kota_n = st.text_input("Nama Kota","Bandung")
    pjg    = st.number_input("Panjang Segmen P (km)",0.1,50.0,1.0,0.1)

    st.subheader("Geometri Jalan")
    tipe = st.selectbox("Tipe Jalan",["2/2-TT","4/2-T","6/2-T"])
    lebar= (st.slider("Lebar Jalur Total LJE (m)",5.0,11.0,7.0,0.5)
            if tipe=="2/2-TT"
            else st.slider("Lebar Per Lajur LLE (m)",3.0,4.0,3.5,0.25))

    ks = st.radio("Kondisi Samping",["Berbahu","Berkereb"])
    if ks=="Berbahu":
        ls=st.slider("Lebar Bahu LBE (m)",0.5,2.0,1.0,0.5); ls_label="LBE"
    else:
        ls=st.slider("Jarak Kereb LKP (m)",0.5,2.0,1.0,0.5); ls_label="LKP"

    hs   = st.selectbox("Kelas Hambatan Samping",["SR","R","S","T","ST"],
                        format_func=lambda x:{"SR":"SR–Sangat Rendah","R":"R–Rendah",
                        "S":"S–Sedang","T":"T–Tinggi","ST":"ST–Sangat Tinggi"}[x])
    kota = st.selectbox("Ukuran Kota",["< 0,1 juta","0,1 – 0,5 juta",
                        "0,5 – 1,0 juta","1,0 – 3,0 juta","> 3,0 juta"])
    split=50
    if tipe=="2/2-TT":
        split=st.slider("Pemisah Arah % dominan",50,70,50,5)

    # ── Jumlah periode (fleksibel 1–14) ──────────────────────────
    st.divider()
    st.subheader("📅 Volume per Periode")
    n_per = st.number_input("Jumlah periode analisis",1,14,3,1)
    st.caption("Total kedua arah" if tipe=="2/2-TT" else "Satu arah dianalisis")

    DEFAULT_JAM = ["07.00–08.00","12.00–13.00","17.00–18.00",
                   "16.00–17.00","18.00–19.00","19.00–20.00",
                   "06.00–07.00","08.00–09.00","09.00–10.00",
                   "10.00–11.00","13.00–14.00","14.00–15.00",
                   "15.00–16.00","20.00–21.00"]
    periodes=[]
    for i in range(int(n_per)):
        with st.expander(f"Periode {i+1}",expanded=(i==0)):
            jam=st.text_input(f"Jam",DEFAULT_JAM[i] if i<len(DEFAULT_JAM) else f"Periode {i+1}",
                              key=f"j{i}")
            kr =st.number_input("KR (MP)",min_value=0,value=500-i*30,key=f"kr{i}")
            kb =st.number_input("KB (KS)",min_value=0,value=50,       key=f"kb{i}")
            sm =st.number_input("SM",     min_value=0,value=800-i*50, key=f"sm{i}")
            periodes.append({"jam":jam,"kr":kr,"kb":kb,"sm":sm})

    hitung_btn = st.button("🔢 Hitung Semua Periode",use_container_width=True)

# ── Hitung & simpan ke session_state ────────────────────────────
if hitung_btn:
    hasil_semua=[hitung(tipe,lebar,ks,ls,hs,kota,split,
                        p["kr"],p["kb"],p["sm"],pjg) for p in periodes]
    st.session_state.data = dict(
        hasil=hasil_semua, periodes=periodes,
        nama=nama, kota_n=kota_n, tipe=tipe, lebar=lebar,
        ks=ks, ls_label=ls_label, ls=ls, hs=hs, kota=kota,
        split=split, pjg=pjg,
    )

# ── Tampilkan hasil (dari session_state agar tidak reset) ────────
D = st.session_state.data
if D is not None:
    hasil_semua = D["hasil"]; periodes = D["periodes"]
    tipe=D["tipe"]; lebar=D["lebar"]; ks=D["ks"]; ls=D["ls"]
    ls_label=D["ls_label"]; hs=D["hs"]; kota=D["kota"]
    split=D["split"]; pjg=D["pjg"]
    nama=D["nama"]; kota_n=D["kota_n"]

    st.subheader(f"📊 Hasil Analisis: {nama}")
    st.caption(f"Kota: {kota_n}  |  Tipe: {tipe}  |  P={pjg} km  |  "
               f"{ks} ({ls_label}={ls} m)  |  KHS: {hs}")

    # Buat tab dinamis
    nama_tabs  = [f"🕐 P{i+1}  {p['jam']}" for i,p in enumerate(periodes)]
    nama_tabs += ["📋 Rekapitulasi"]
    tabs = st.tabs(nama_tabs)

    # Tab per periode
    for i,(tab,h,p) in enumerate(zip(tabs[:-1],hasil_semua,periodes)):
        with tab:
            tampil_periode(h,p,i,tipe,lebar,ks,ls_label,ls,hs,kota,split,pjg,nama)

    # Tab rekapitulasi
    with tabs[-1]:
        st.subheader("📋 Rekapitulasi Semua Periode")
        df_r=pd.DataFrame({
            "Periode":     [p["jam"] for p in periodes],
            "Q (smp/jam)": [f"{h['q']:.1f}"  for h in hasil_semua],
            "C (smp/jam)": [f"{h['c']:.1f}"  for h in hasil_semua],
            "vB (km/jam)": [f"{h['vb']:.1f}" for h in hasil_semua],
            "DJ":          [f"{h['dj']:.3f}"  for h in hasil_semua],
            "LOS":         [h["los"]           for h in hasil_semua],
            "vT (km/jam)": [f"{h['vt']:.1f}"  for h in hasil_semua],
            "wT (menit)":  [f"{h['wt']:.2f}"  for h in hasil_semua],
            "Deskripsi":   [h["desk"]          for h in hasil_semua],
        })
        st.dataframe(df_r,use_container_width=True,hide_index=True)
        st.divider()

        # Gauge berdampingan (maks 3 per baris)
        st.subheader("Gauge DJ — Semua Periode")
        for row_start in range(0,len(hasil_semua),3):
            cols=st.columns(min(3,len(hasil_semua)-row_start))
            for col,h,p in zip(cols,hasil_semua[row_start:row_start+3],
                                periodes[row_start:row_start+3]):
                with col:
                    st.plotly_chart(gauge(h["dj"],p["jam"]),
                                    use_container_width=True,
                                    key=f"gr_{row_start}_{p['jam']}")
        st.divider()

        # Bar DJ
        clr=["#27ae60" if h["dj"]<=0.44 else "#f39c12" if h["dj"]<=0.82
             else "#e74c3c" for h in hasil_semua]
        fig=go.Figure()
        fig.add_trace(go.Bar(
            x=[p["jam"] for p in periodes],
            y=[h["dj"] for h in hasil_semua],
            marker_color=clr,
            text=[f"DJ={h['dj']:.3f}<br>LOS {h['los']}" for h in hasil_semua],
            textposition="outside"))
        for bts,lbl_g,wg in [(.44,"B/C","#27ae60"),(.64,"C/D","#d4ac0d"),
                              (.82,"D/E","#e67e22"),(1.00,"E/F","#e74c3c")]:
            fig.add_hline(y=bts,line_dash="dot",line_color=wg,
                          annotation_text=f"Batas {lbl_g}",
                          annotation_position="right",annotation_font_color=wg)
        fig.update_layout(
            title="Perbandingan DJ antar Periode",
            yaxis=dict(range=[0,max(h["dj"] for h in hasil_semua)*1.4+0.1],title="DJ"),
            xaxis_title="Periode",height=380,showlegend=False,
            paper_bgcolor="rgba(0,0,0,0)",plot_bgcolor="rgba(0,0,0,0)",
            font=dict(color="white"))
        st.plotly_chart(fig,use_container_width=True,key="bar_rekap")

        # Q vs C
        st.subheader("Volume Q vs Kapasitas C")
        st.bar_chart(pd.DataFrame({
            "Periode":    [p["jam"] for p in periodes],
            "Q (smp/jam)":[round(h["q"],1) for h in hasil_semua],
            "C (smp/jam)":[round(h["c"],1) for h in hasil_semua],
        }).set_index("Periode"))
        st.divider()

        # ── Download Excel ────────────────────────────────────────
        st.subheader("💾 Download Laporan")
        xl=buat_excel(nama,kota_n,tipe,pjg,ks,ls_label,ls,hs,kota,periodes,hasil_semua)
        st.download_button(
            "📊 Download Laporan Excel (.xlsx)",
            data=xl,
            file_name=f"{nama.replace(' ','_')}_PKJI2023.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
else:
    st.info("👈 Isi data di sidebar, lalu klik **Hitung Semua Periode**")
